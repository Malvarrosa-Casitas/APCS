
from flask import Flask, request, jsonify, send_file, session
from flask_cors import CORS
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from pathlib import Path
import sqlite3, base64, datetime, os, smtplib, secrets
from functools import wraps
from email.message import EmailMessage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
BASE=Path(__file__).resolve().parents[1]
DB=BASE/'database.db'
app=Flask(__name__, static_folder=str(BASE), static_url_path='')
app.secret_key=os.getenv('SECRET_KEY','dev-change-this-secret-key')
app.config.update(SESSION_COOKIE_HTTPONLY=True, SESSION_COOKIE_SAMESITE='Lax', PERMANENT_SESSION_LIFETIME=datetime.timedelta(minutes=int(os.getenv('SESSION_MINUTES','30'))))
CORS(app, supports_credentials=True)
ADMIN_USERNAME=os.getenv('ADMIN_USERNAME','admin')
ADMIN_PASSWORD_HASH=os.getenv('ADMIN_PASSWORD_HASH') or generate_password_hash(os.getenv('ADMIN_PASSWORD','casitasrosas2026'))
FAILED_LOG=BASE/'admin_failed_logins.log'
OFFICIAL={'Carrer Beat Juan Grande':[16,18,20,22,24,26,28],'Carrer Pare Pedro Velasco':[1,2,3,4,5,6,7,8,9,10,11,12,13],'Carrer Pare Antón Martín':[27,29,31,33,35,37,46,48,50,52,54,56],'Avinguda Malva-Rosa':[43,45,47,49,51,53],'Carrer Sant Rafael':[24,26,28,30,32,34],'Carrer Sant Joan De Déu':[25,29,31,33,35,37]}
COLLAB=['Miembro activo','Seguridad','Organizar actividades','Informar incidencias','Seguimiento escritos','Aportar ideas','Otros']

def con():
    c=sqlite3.connect(DB); c.row_factory=sqlite3.Row; return c

def ensure_column(table, column, definition):
    with con() as c:
        cols=[r['name'] for r in c.execute(f'PRAGMA table_info({table})')]
        if column not in cols: c.execute(f'ALTER TABLE {table} ADD COLUMN {column} {definition}')

def init_db():
    with con() as c:
        c.executescript("""
        CREATE TABLE IF NOT EXISTS registros(id INTEGER PRIMARY KEY AUTOINCREMENT,tipo TEXT NOT NULL CHECK(tipo IN ('socio','informado')),nombre TEXT NOT NULL,direccion TEXT NOT NULL,telefono TEXT NOT NULL,email TEXT,unidad_familiar TEXT,propietario TEXT,portal TEXT,puerta TEXT,quiere_participar TEXT,tipos_colaboracion TEXT,ideas TEXT,problemas TEXT,firma_path TEXT,fecha_alta TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS incidencias(id INTEGER PRIMARY KEY AUTOINCREMENT,nombre TEXT NOT NULL,apellidos TEXT NOT NULL,direccion TEXT,telefono TEXT NOT NULL,descripcion TEXT,incluir_instituciones TEXT,fecha TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS fotografias(id INTEGER PRIMARY KEY AUTOINCREMENT,tipo TEXT,ref_id INTEGER,path TEXT,fecha TEXT NOT NULL);
        """)
    for table, cols in {'registros':{'unidad_familiar':'TEXT','propietario':'TEXT','quiere_participar':'TEXT','tipos_colaboracion':'TEXT'}, 'incidencias':{'incluir_instituciones':'TEXT'}}.items():
        for col, definition in cols.items(): ensure_column(table, col, definition)

def now_iso(): return datetime.datetime.now().isoformat(timespec='seconds')

def refresh_activity():
    session.permanent=True; session['last_activity']=datetime.datetime.utcnow().isoformat()

def session_valid():
    if not session.get('admin_authenticated'): return False
    last=session.get('last_activity')
    if not last: return False
    try: delta=datetime.datetime.utcnow()-datetime.datetime.fromisoformat(last)
    except Exception: return False
    if delta > app.config['PERMANENT_SESSION_LIFETIME']:
        session.clear(); return False
    refresh_activity(); return True

def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session_valid(): return jsonify({'error':'admin authentication required'}), 401
        return fn(*args, **kwargs)
    return wrapper

def csrf_token():
    token=session.get('csrf_token')
    if not token:
        token=secrets.token_urlsafe(32); session['csrf_token']=token
    return token

def csrf_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        token=request.form.get('csrf_token') or request.headers.get('X-CSRFToken')
        if not token or token != session.get('csrf_token'):
            return jsonify({'error':'invalid csrf token'}), 400
        return fn(*args, **kwargs)
    return wrapper

def log_failed(username):
    FAILED_LOG.write_text((FAILED_LOG.read_text(encoding='utf-8') if FAILED_LOG.exists() else '') + f"{now_iso()}\t{username}\t{request.remote_addr}\n", encoding='utf-8')

def save_files(files,tipo,ref_id):
    folder=BASE/'uploads'/'photos'; folder.mkdir(parents=True,exist_ok=True)
    for f in files:
        if not f or not f.filename: continue
        name=f'{tipo}_{ref_id}_{secure_filename(f.filename)}'; path=folder/name; f.save(path)
        with con() as c: c.execute('INSERT INTO fotografias(tipo,ref_id,path,fecha) VALUES(?,?,?,?)',(tipo,ref_id,str(path.relative_to(BASE)),now_iso()))

def save_signature(data,ref_id,tipo):
    if not data or ',' not in data: return ''
    folder=BASE/'uploads'/'firmas'; folder.mkdir(parents=True,exist_ok=True)
    path=folder/f'firma_{tipo}_{ref_id}.png'; path.write_bytes(base64.b64decode(data.split(',',1)[1])); return str(path.relative_to(BASE))

def send_confirmation(to_addr,tipo,nombre):
    if not to_addr: return False
    host=os.getenv('SMTP_HOST'); user=os.getenv('SMTP_USER'); pwd=os.getenv('SMTP_PASSWORD'); sender=os.getenv('SMTP_FROM',user or '')
    if not host or not sender: return False
    msg=EmailMessage(); msg['Subject']='Confirmación Las Casitas Rosas'; msg['From']=sender; msg['To']=to_addr
    msg.set_content(f'Hola {nombre or ""}, hemos recibido tu formulario de {tipo}. Gracias por participar en Las Casitas Rosas.')
    with smtplib.SMTP(host,int(os.getenv('SMTP_PORT','587'))) as s:
        s.starttls()
        if user and pwd: s.login(user,pwd)
        s.send_message(msg)
    return True

def parse_portal(portal):
    if not portal or ':' not in portal: return None,None
    calle,num=portal.rsplit(':',1)
    try: n=int(num.strip())
    except: return calle.strip(), None
    return calle.strip(), n

def street_representation(c):
    represented={street:set() for street in OFFICIAL}
    for r in c.execute("SELECT DISTINCT portal FROM registros WHERE propietario='si' AND portal IS NOT NULL AND trim(portal)!=''"):
        street,n=parse_portal(r['portal'])
        if street in OFFICIAL and n in OFFICIAL[street]: represented[street].add(n)
    rows=[]
    for street, nums in OFFICIAL.items():
        pct=round(len(represented[street])*100/len(nums),1)
        rows.append((street,pct))
    rows.sort(key=lambda x:x[1], reverse=True)
    return rows

@app.route('/')
def home(): return app.send_static_file('index.html')

@app.get('/api/csrf-token')
def get_csrf(): return jsonify({'csrf_token':csrf_token()})

@app.post('/api/admin/login')
@csrf_required
def admin_login():
    username=request.form.get('username','').strip(); password=request.form.get('password','')
    if username==ADMIN_USERNAME and check_password_hash(ADMIN_PASSWORD_HASH,password):
        session['admin_authenticated']=True; session['admin_username']=username; refresh_activity(); csrf_token(); return jsonify({'ok':True})
    log_failed(username); return jsonify({'ok':False}), 401

@app.post('/api/admin/logout')
@csrf_required
def admin_logout():
    session.clear(); csrf_token(); return jsonify({'ok':True})

@app.get('/api/admin/status')
def admin_status(): return jsonify({'authenticated':session_valid()})

@app.post('/api/registros')
@csrf_required
def registros():
    init_db(); f=request.form; tipo=f.get('tipo','informado'); fecha=now_iso()
    if tipo not in ('socio','informado'): tipo='informado'
    tipos=','.join(request.form.getlist('tipos_colaboracion')) if f.get('quiere_participar')=='si' else ''
    with con() as c:
        cur=c.execute("""INSERT INTO registros(tipo,nombre,direccion,telefono,email,unidad_familiar,propietario,portal,puerta,quiere_participar,tipos_colaboracion,ideas,problemas,firma_path,fecha_alta) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",(tipo,f.get('nombre'),f.get('direccion'),f.get('telefono'),f.get('email'),f.get('unidad_familiar'),f.get('propietario'),f.get('portal'),f.get('puerta'),f.get('quiere_participar'),tipos,f.get('ideas'),f.get('problemas'),' ',fecha))
        rid=cur.lastrowid
    sig=save_signature(f.get('firma_png'),rid,tipo); save_files(request.files.getlist('photos'),tipo,rid)
    with con() as c: c.execute('UPDATE registros SET firma_path=? WHERE id=?',(sig,rid))
    return jsonify({'ok':True,'id':rid,'email_sent':send_confirmation(f.get('email'),tipo,f.get('nombre'))})

@app.post('/api/incidencias')
@csrf_required
def incidencias():
    init_db(); f=request.form; fecha=now_iso(); incluir='si' if f.get('incluir_instituciones')=='si' else 'no'
    with con() as c:
        cur=c.execute('INSERT INTO incidencias(nombre,apellidos,direccion,telefono,descripcion,incluir_instituciones,fecha) VALUES(?,?,?,?,?,?,?)',(f.get('nombre'),f.get('apellidos'),f.get('direccion'),f.get('telefono'),f.get('descripcion'),incluir,fecha)); iid=cur.lastrowid
    save_files(request.files.getlist('photos'),'incidencia',iid)
    return jsonify({'ok':True,'id':iid,'email_sent':False})

@app.get('/api/registros')
@admin_required
def get_registros():
    init_db()
    with con() as c: rows=[dict(r) for r in c.execute('SELECT id,tipo,nombre,telefono,propietario,portal,quiere_participar,fecha_alta FROM registros ORDER BY id DESC LIMIT 300')]
    return jsonify(rows)

@app.get('/api/stats')
def stats():
    init_db()
    with con() as c:
        socios=c.execute("SELECT COUNT(*) n FROM registros WHERE tipo='socio'").fetchone()['n']
        informados=c.execute("SELECT COUNT(*) n FROM registros WHERE tipo='informado'").fetchone()['n']
        ideas=c.execute("SELECT COUNT(*) n FROM registros WHERE ideas IS NOT NULL AND trim(ideas)!=''").fetchone()['n']
        incidencias=c.execute('SELECT COUNT(*) n FROM incidencias').fetchone()['n']
        fotografias=c.execute('SELECT COUNT(*) n FROM fotografias').fetchone()['n']
        participacion=c.execute("SELECT COUNT(*) n FROM registros WHERE quiere_participar='si'").fetchone()['n']
        ui=c.execute('SELECT nombre,fecha_alta FROM registros ORDER BY fecha_alta DESC LIMIT 1').fetchone()
        uinc=c.execute('SELECT nombre,fecha FROM incidencias ORDER BY fecha DESC LIMIT 1').fetchone()
        reps=street_representation(c)
        coll_counts={k:0 for k in COLLAB}
        for r in c.execute("SELECT tipos_colaboracion FROM registros WHERE tipos_colaboracion IS NOT NULL AND trim(tipos_colaboracion)!=''"):
            for item in [x.strip() for x in r['tipos_colaboracion'].split(',') if x.strip()]:
                if item in coll_counts: coll_counts[item]+=1
    return jsonify({'socios':socios,'informados':informados,'ideas':ideas,'incidencias':incidencias,'fotografias':fotografias,'participacion':participacion,'ultima_inscripcion':(ui['fecha_alta'][:10] if ui else '—'),'ultima_incidencia':(uinc['fecha'][:10] if uinc else '—'),'representacion_calles_labels':[r[0] for r in reps],'representacion_calles_valores':[r[1] for r in reps],'colaboracion_labels':list(coll_counts.keys()),'colaboracion_valores':list(coll_counts.values())})

@app.get('/api/export/registros')
@admin_required
def export_registros():
    init_db(); out=BASE/'exports'/'Registros_Casitas_Rosas.xlsx'; out.parent.mkdir(parents=True,exist_ok=True); wb=Workbook()
    def style(ws):
        for cell in ws[1]: cell.font=Font(bold=True,color='FFFFFF'); cell.fill=PatternFill('solid',fgColor='A87CC4'); cell.alignment=Alignment(horizontal='center')
        ws.auto_filter.ref=ws.dimensions
        for col in ws.columns: ws.column_dimensions[get_column_letter(col[0].column)].width=min(max(len(str(c.value or '')) for c in col)+2,50)
    ws=wb.active; ws.title='Registros'; headers=['id','tipo','nombre','direccion','telefono','email','unidad_familiar','propietario','portal','puerta','quiere_participar','tipos_colaboracion','ideas','problemas','firma_path','fecha_alta']; ws.append(headers)
    with con() as c:
        for r in c.execute('SELECT * FROM registros ORDER BY id'): ws.append([r[h] for h in headers])
    style(ws)
    ws2=wb.create_sheet('Incidencias'); h2=['id','nombre','apellidos','direccion','telefono','descripcion','incluir_instituciones','fecha']; ws2.append(h2)
    with con() as c:
        for r in c.execute('SELECT * FROM incidencias ORDER BY id'): ws2.append([r[h] for h in h2])
    style(ws2)
    ws3=wb.create_sheet('Fotografias'); h3=['id','tipo','ref_id','path','fecha']; ws3.append(h3)
    with con() as c:
        for r in c.execute('SELECT * FROM fotografias ORDER BY id'): ws3.append([r[h] for h in h3])
    style(ws3); wb.save(out)
    return send_file(out,as_attachment=True,download_name='Registros_Casitas_Rosas.xlsx')

if __name__=='__main__': init_db(); app.run(debug=True)
